from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from models.receitas import ReceitaModel
from models.despesas import DespesaModel


class ExportadorExcel:

    def __init__(self):
        self.receita_model = ReceitaModel()
        self.despesa_model = DespesaModel()

    def exportar(self, caminho_arquivo):

        wb = Workbook()
        wb.remove(wb.active)

        # Aba de Receitas
        self.adicionar_aba_receitas(wb)

        # Aba de Despesas
        self.adicionar_aba_despesas(wb)

        # Aba de Resumo
        self.adicionar_aba_resumo(wb)

        # Salvar arquivo
        wb.save(caminho_arquivo)

        return True

    def adicionar_aba_receitas(self, wb):

        ws = wb.create_sheet("Receitas")

        # Cabeçalhos
        cabecalhos = ["ID", "Nome", "Valor"]
        ws.append(cabecalhos)

        # Estilo dos cabeçalhos
        estilo_cabecalho = self.obter_estilo_cabecalho()
        for cell in ws[1]:
            cell.font = estilo_cabecalho["font"]
            cell.fill = estilo_cabecalho["fill"]
            cell.alignment = estilo_cabecalho["alignment"]
            cell.border = estilo_cabecalho["border"]

        # Dados
        receitas = self.receita_model.listar()
        for receita in receitas:
            ws.append([receita[0], receita[1], f"R$ {receita[2]:.2f}"])

        # Ajustar largura das colunas
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15

        # Formatar células de valor
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")

    def adicionar_aba_despesas(self, wb):

        ws = wb.create_sheet("Despesas")

        # Cabeçalhos
        cabecalhos = ["ID", "Data", "Mês", "Nome", "Valor"]
        ws.append(cabecalhos)

        # Estilo dos cabeçalhos
        estilo_cabecalho = self.obter_estilo_cabecalho()
        for cell in ws[1]:
            cell.font = estilo_cabecalho["font"]
            cell.fill = estilo_cabecalho["fill"]
            cell.alignment = estilo_cabecalho["alignment"]
            cell.border = estilo_cabecalho["border"]

        # Dados
        despesas = self.despesa_model.listar()
        for despesa in despesas:
            ws.append(
                [despesa[0], despesa[1], despesa[2], despesa[3], f"R$ {despesa[4]:.2f}"]
            )

        # Ajustar largura das colunas
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 15

        # Formatar células de valor
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")

    def adicionar_aba_resumo(self, wb):

        ws = wb.create_sheet("Resumo", 0)

        # Título
        ws.merge_cells("A1:B1")
        titulo = ws["A1"]
        titulo.value = "RESUMO FINANCEIRO"
        titulo.font = Font(size=16, bold=True, color="FFFFFF")
        titulo.fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Dados
        total_receitas = self.receita_model.total_receitas()
        total_despesas = self.despesa_model.total_despesas()
        saldo = total_receitas - total_despesas
        data_exportacao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Linha 3: Total Receitas
        ws["A3"] = "Total de Receitas:"
        ws["B3"] = f"R$ {total_receitas:.2f}"
        self.aplicar_estilo_linha(ws, 3, background="C6E0B4")

        # Linha 4: Total Despesas
        ws["A4"] = "Total de Despesas:"
        ws["B4"] = f"R$ {total_despesas:.2f}"
        self.aplicar_estilo_linha(ws, 4, background="F4B084")

        # Linha 5: Saldo
        ws["A5"] = "Saldo:"
        ws["B5"] = f"R$ {saldo:.2f}"
        cor_saldo = "A9D08E" if saldo >= 0 else "F8CBAD"
        self.aplicar_estilo_linha(ws, 5, background=cor_saldo)

        # Linha 7: Data de Exportação
        ws["A7"] = "Data de Exportação:"
        ws["B7"] = data_exportacao

        # Ajustar largura
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

    def obter_estilo_cabecalho(self):

        return {
            "font": Font(size=12, bold=True, color="FFFFFF"),
            "fill": PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            ),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    def aplicar_estilo_linha(self, ws, linha, background="FFFFFF"):

        for col in ["A", "B"]:
            cell = ws[f"{col}{linha}"]
            cell.fill = PatternFill(
                start_color=background, end_color=background, fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
